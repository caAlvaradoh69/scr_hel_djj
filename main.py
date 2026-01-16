import io
import re
import json
import time
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional

from playwright.sync_api import sync_playwright, TimeoutError
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from googleapiclient.http import MediaFileUpload
import os
from datetime import datetime
from zoneinfo import ZoneInfo

# ======================================================
# CONFIG
# ======================================================
SERVICE_ACCOUNT_JSON = "price-scraper-drive-bafdf4e1b138.json"
DRIVE_FILE_ID = "16dxM-FoL86i8bUsQ_DKoFpq5vVE9PepZ"  # <-- CAMBIA ESTO
DRIVE_UPLOAD_FOLDER_ID = "15BGs93qdpZbhGg3NrCokBQMNuCbt5B7K"
FECHA = datetime.now(ZoneInfo("America/Santiago")).strftime("%d-%m-%Y %H:%M")
HEADLESS = True                 # recomendado False
PAGE_TIMEOUT = 60000
DELAY = 1.5

OUT_FILE = "output_prices_djichile.json"

# ======================================================
# HELPERS
# ======================================================
def parse_price(text: str) -> Optional[int]:
    if not text:
        return None
    digits = re.sub(r"[^\d]", "", text)
    return int(digits) if digits else None

# ======================================================
# GOOGLE DRIVE
# ======================================================
def download_excel() -> io.BytesIO:
    scopes = ["https://www.googleapis.com/auth/drive.readonly"]
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_JSON, scopes=scopes
    )
    svc = build("drive", "v3", credentials=creds)
    req = svc.files().get_media(fileId=DRIVE_FILE_ID)

    fh = io.BytesIO()
    MediaIoBaseDownload(fh, req).next_chunk()
    fh.seek(0)
    return fh

def upload_excel_to_drive(
    file_path: str,
    folder_id: str,
    drive_filename: Optional[str] = None
):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"No existe el archivo: {file_path}")

    delete_previous_excels_from_drive(
        folder_id=DRIVE_UPLOAD_FOLDER_ID,
        filename_prefix="precios_djjchile"
    )
    scopes = ["https://www.googleapis.com/auth/drive.file"]
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_JSON, scopes=scopes
    )

    service = build("drive", "v3", credentials=creds)

    file_metadata = {
        "name": drive_filename or os.path.basename(file_path),
        "parents": [folder_id],
    }

    media = MediaFileUpload(
        file_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )

    uploaded_file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, webViewLink",
        supportsAllDrives=True
    ).execute()

    print("☁️ Excel subido a Drive")
    print("🔗 Link:", uploaded_file["webViewLink"])

    return uploaded_file

def delete_previous_excels_from_drive(
    folder_id: str,
    filename_prefix: str
):
    """
    Elimina archivos Excel en Drive dentro de una carpeta
    cuyo nombre comience con filename_prefix
    """

    scopes = ["https://www.googleapis.com/auth/drive"]
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_JSON, scopes=scopes
    )
    service = build("drive", "v3", credentials=creds)

    query = (
        f"'{folder_id}' in parents "
        f"and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
        f"and name contains '{filename_prefix}' "
        f"and trashed = false"
    )

    results = service.files().list(
        q=query,
        fields="files(id, name)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True
    ).execute()

    files = results.get("files", [])

    if not files:
        print("🟢 No hay archivos antiguos que eliminar")
        return

    for f in files:
        service.files().delete(
            fileId=f["id"],
            supportsAllDrives=True
        ).execute()
        print(f"🗑️ Eliminado de Drive: {f['name']}")
# ======================================================
# EXCEL
# ======================================================
def load_products(xlsx: io.BytesIO) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active

    headers = {
        str(ws.cell(1, c).value).lower(): c
        for c in range(1, ws.max_column + 1)
    }

    required = ["sku", "nombre_producto", "precio_publico", "link"]
    for r in required:
        if r not in headers:
            raise RuntimeError(f"Falta columna requerida: {r}")

    products = []
    for r in range(2, ws.max_row + 1):
        url = ws.cell(r, headers["link"]).value
        if not url:
            continue

        products.append({
            "sku": str(ws.cell(r, headers["sku"]).value).strip(),
            "product_name": str(ws.cell(r, headers["nombre_producto"]).value).strip(),
            "base_price": ws.cell(r, headers["precio_publico"]).value,
            "url": str(url).strip()
        })

    return products

# ======================================================
# SCRAPING djichile (PRECIO)
# ======================================================
def extract_price_from_page(page) -> Optional[int]:
    try:
        page.wait_for_timeout(2000)

        price = page.evaluate("""
        () => {
            // 1️⃣ OpenGraph (FUENTE PRINCIPAL)
            const og = document.querySelector(
                'meta[property="product:price:amount"]'
            );
            if (og && og.content) {
                return og.content;
            }

            // 2️⃣ Microdata fallback
            const micro = document.querySelector('meta[itemprop="price"]');
            if (micro && micro.content) {
                return micro.content;
            }

            // 3️⃣ Visible fallback
            const spans = Array.from(document.querySelectorAll("span"));
            for (const s of spans) {
                const t = s.textContent?.replace(/\\u00a0/g, " ").trim();
                if (t && t.startsWith("$") && t.length < 18) {
                    return t;
                }
            }

            return null;
        }
        """)

        if not price:
            return None

        digits = re.sub(r"[^\d]", "", str(price))
        return int(digits) if digits else None

    except Exception as e:
        print("❌ Error extrayendo precio:", e)
        return None

def export_to_excel(results, output_file="output_prices_djichile.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios djichile"

    headers = [
        "sku",
        "nombre_producto",
        "precio_publico",
        "precio_djichile",
        "diferencia",
        "url"
    ]

    ws.append(headers)

    highlight = PatternFill(
        start_color="FFF59D",  # amarillo suave
        end_color="FFF59D",
        fill_type="solid"
    )

    for r in results:
        precio_publico = r.get("base_price")
        precio_djichile = r.get("price")

        diferencia = None
        if precio_publico is not None and precio_djichile is not None:
            diferencia = precio_publico - precio_djichile

        row = [
            r.get("sku"),
            r.get("product_name"),
            precio_publico,        # ✅ PRECIO BASE CORRECTO
            precio_djichile,
            diferencia,
            r.get("url")
        ]

        ws.append(row)
        row_idx = ws.max_row

        # 🟨 Marcar oportunidad real
        if (
            precio_publico is not None
            and precio_djichile is not None
            and precio_djichile < precio_publico
        ):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = highlight

    wb.save(output_file)
    upload_excel_to_drive(output_file, DRIVE_UPLOAD_FOLDER_ID, f"precios_djjchile_{FECHA}.xlsx")
    print(f"📊 Excel generado correctamente: {output_file}")
# ======================================================
# MAIN
# ======================================================
def main():
    print("⬇ Descargando Excel desde Drive...")
    xlsx = download_excel()

    print("📖 Cargando productos...")
    products = load_products(xlsx)
    print(f"📦 Productos a procesar: {len(products)}")

    results = []
    now = datetime.utcnow().isoformat()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(locale="es-CL")
        page = context.new_page()

        # 🚀 bloquear recursos pesados
        page.route("*/", lambda r, q:
            r.abort() if q.resource_type in ("image", "font", "media") else r.continue_()
        )

        for prod in products:
            print(f"\n🔎 Procesando: {prod['product_name']}")
            print(f"🔗 URL: {prod['url']}")

            try:
                page.goto(
                    prod["url"],
                    wait_until="domcontentloaded",
                    timeout=PAGE_TIMEOUT
                )

                page.wait_for_timeout(2500)
                price = extract_price_from_page(page)

                print("💰 PRECIO FINAL:", price)

            except Exception as e:
                print("❌ Error navegando:", e)
                price = None

            results.append({
                "sku": prod["sku"],
                "product_name": prod["product_name"],
                "base_price":prod["base_price"],
                "price": price,
                "process_date": now,
                "source": "djichile",
                "url": prod["url"]
            })

            time.sleep(DELAY)

        browser.close()
        export_to_excel(results)

    # print("\n📤 RESULTADO FINAL:")
    # print(json.dumps(results, indent=2, ensure_ascii=False))

    # with open(OUT_FILE, "w", encoding="utf-8") as f:
    #     json.dump(results, f, indent=2, ensure_ascii=False)

    # print(f"\n✅ Guardado en {OUT_FILE}")

if __name__ == "__main__":
    main()